from openpyxl import load_workbook, workbook

arquivo_alunos = load_workbook('alunos.xlsx')
planilha_alunos = arquivo_alunos['Alunos']

alunos_aprovados = []
alunos_reprovados = []

for aluno in planilha_alunos.iter_rows(min_row=2, values_only=True):
    nome, curso, idade, nota_final, data_matricula = aluno
    aluno_dicionario = {
        'Nome': nome,
        'Curso': curso,
        'Idade': idade,
        'Nota Final': nota_final,
        'Data Matrícula': data_matricula
    }
    if nota_final >= 7:
        alunos_aprovados.append(aluno_dicionario)
    else:
        alunos_reprovados.append(aluno_dicionario)
 
arquivo_aprovados = workbook()
planilha_aprovados = arquivo_aprovados.active
planilha_aprovados.title = 'Alunos Aprovados'

arquivo_reprovados = workbook()
planilha_reprovados = arquivo_reprovados.active
planilha_reprovados.title = 'Alunos reprovados'


arquivo_aprovados.save('aprovados.xlsx')
arquivo_reprovados.save('reprovados.xlsx')

