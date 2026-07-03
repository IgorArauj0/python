from openpyxl import Workbook

arquivo = Workbook()

planilha_atual = arquivo.active
planilha_atual.title = "Produtos"

# planilha_atual["A1"] = "Produto"
# planilha_atual["B1"] = "Preço"

# planilha_atual["A2"] = "Camiseta"
# planilha_atual["B2"] = 29.90

# planilha_atual["A3"] = "Calça"
# planilha_atual["B3"] = 79.90

# planilha_atual["A4"] = "Tênis"
# planilha_atual["B4"] = 199.90

planilha_atual.append(["Produto", "Preço"])
planilha_atual.append(["Camiseta", 29.90])
planilha_atual.append(["Calça", 79.90])
planilha_atual.append(["Tênis", 199.90])
planilha_atual.append(["Meia", 9.90])

planilha_atual['B4'] = 89.90


planilha_vendas = arquivo.create_sheet("Vendas")
planilha_vendas.append(["Produto", "Quantidade", "Preço Unitário", "Total"])
planilha_vendas.append(["Camiseta", 2, 29.90, 2 * 29.90])
planilha_vendas.append(["Calça", 1, 79.90, 1 * 79.90])
planilha_vendas.append(["Tênis", 1, 199.90, 1 * 199.90])
planilha_vendas.append(["Meia", 5, 9.90, 5 * 9.90])

print(arquivo.sheetnames)

arquivo.save("planilha.xlsx")