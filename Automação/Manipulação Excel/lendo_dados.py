# Importa o módulo 'sys', responsável por fornecer acesso a recursos internos
# do interpretador Python, como entrada (stdin), saída (stdout) e erros (stderr).
import sys

#Sempre que você enviar texto para o terminal, utilize a codificação UTF-8.
sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import load_workbook

arquivo = load_workbook("planilha_funcionarios.xlsx")

planilha_funcionarios = arquivo['Funcionários']

# salario_antonio = planilha_funcionarios['D6'].value
# print(f"Salário do Antonio: {salario_antonio}")

# linha_13 = planilha_funcionarios[13]
# for celula in linha_13:
#     print(celula.value)

# linhas = planilha_funcionarios[7:12]
# for linha in linhas:
#     print("="*30)
#     for celula in linha:
#         print(celula.value)

# coluna_salario = planilha_funcionarios['D']
# for celula in coluna_salario:
#     print(celula.value)

# for linha in planilha_funcionarios.iter_rows(values_only=True):
#     print("="*30)
#     for celula in linha:
#         print(celula)

for linha in planilha_funcionarios.iter_rows(values_only=True, min_row=1, max_row=20):
    print('-'*50)
    #desempacotamento de valores
    nome, departamento, idade, salario, data_admissao = linha
    
    # Verificar o tipo de dado da data
    # Se data_admissao for uma string (texto), usa-a diretamente
    # Se for um objeto datetime, formata com strftime
    if isinstance(data_admissao, str):
        data_formatada = data_admissao
    else:
        # Se for um objeto datetime, converte para o formato DD/MM/YYYY
        data_formatada = data_admissao.strftime("%d/%m/%Y")
    
    print(f"""
    Nome: {nome}
    Departamento: {departamento}
    Idade: {idade}
    Salário: {salario}
    Data de Admissão: {data_formatada}
    """)